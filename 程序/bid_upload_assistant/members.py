from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Any


TYPE_CODES = {
    "项目负责人/总监": "1",
    "技术负责人": "3",
    "安全负责人": "4",
    "其他": "9",
}
HEADERS = ["成员姓名", "成员类型", "成员身份证号", "职业资格名称", "资格证书编号"]
ALIASES = {
    "成员姓名": "bid_member_name", "姓名": "bid_member_name", "bid_member_name": "bid_member_name",
    "成员类型": "bid_member_type", "成员类型代码": "bid_member_type", "bid_member_type": "bid_member_type",
    "成员身份证号": "bid_member_cert_no", "身份证号": "bid_member_cert_no", "bid_member_cert_no": "bid_member_cert_no",
    "职业资格名称": "bid_member_certificate_name", "bid_member_certificate_name": "bid_member_certificate_name",
    "资格证书编号": "bid_member_certificate_num", "bid_member_certificate_num": "bid_member_certificate_num",
}
REQUIRED = ("bid_member_name", "bid_member_type", "bid_member_cert_no")


def write_template(target: Path) -> Path:
    target.parent.mkdir(parents=True, exist_ok=True)
    with target.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(HEADERS)
        writer.writerow(["示例姓名", "项目负责人/总监", "示例身份证号", "示例资格名称", "示例资格证书编号"])
    return target


def load_members(source: Path) -> list[dict[str, str]]:
    source = Path(source)
    if source.suffix.lower() == ".csv":
        with source.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.DictReader(handle))
    elif source.suffix.lower() in {".xlsx", ".xls"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("读取 Excel 需要 openpyxl") from exc
        workbook = load_workbook(source, read_only=True, data_only=True)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            return []
        headers = [str(value or "").strip() for value in values[0]]
        rows = [dict(zip(headers, ("" if value is None else str(value) for value in row))) for row in values[1:]]
    else:
        raise ValueError("成员资料仅支持 CSV、XLSX 或 XLS")
    result: list[dict[str, str]] = []
    for row in rows:
        mapped: dict[str, str] = {}
        for column, value in row.items():
            key = ALIASES.get(str(column).strip())
            if key:
                mapped[key] = str(value or "").strip()
        if any(mapped.values()):
            result.append(_normalize_member(mapped))
    validate_members(result)
    return result


def _normalize_member(row: dict[str, str]) -> dict[str, str]:
    item = {key: str(row.get(key, "")).strip() for key in ALIASES.values()}
    item["bid_member_type"] = TYPE_CODES.get(item["bid_member_type"], item["bid_member_type"])
    return {key: item[key] for key in ("bid_member_name", "bid_member_type", "bid_member_cert_no", "bid_member_certificate_name", "bid_member_certificate_num")}


def validate_members(members: list[dict[str, str]]) -> None:
    if not members:
        raise ValueError("成员表没有有效数据")
    seen: set[tuple[str, str]] = set()
    for index, member in enumerate(members, 1):
        missing = [key for key in REQUIRED if not member.get(key)]
        if missing:
            raise ValueError(f"第 {index} 行缺少必填字段: {', '.join(missing)}")
        if member["bid_member_type"] not in set(TYPE_CODES.values()):
            raise ValueError(f"第 {index} 行成员类型必须是：{'、'.join(TYPE_CODES)}")
        cert = member["bid_member_cert_no"].upper()
        if not re.fullmatch(r"\d{17}[\dX]", cert):
            raise ValueError(f"第 {index} 行身份证号格式不正确")
        member["bid_member_cert_no"] = cert
        if member["bid_member_type"] == "1" and (not member["bid_member_certificate_name"] or not member["bid_member_certificate_num"]):
            raise ValueError(f"第 {index} 行项目负责人/总监必须填写职业资格名称和资格证书编号")
        marker = (member["bid_member_type"], cert)
        if marker in seen:
            raise ValueError(f"第 {index} 行成员类型和身份证号与前面重复")
        seen.add(marker)


def equivalent(left: list[dict[str, Any]], right: list[dict[str, Any]]) -> bool:
    keys = ("bid_member_name", "bid_member_type", "bid_member_cert_no", "bid_member_certificate_name", "bid_member_certificate_num")
    def normalize(items: list[dict[str, Any]]) -> list[tuple[str, ...]]:
        return sorted(tuple(str(item.get(key, "")).strip() for key in keys) for item in items)
    return normalize(left) == normalize(right)

